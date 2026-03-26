#!/usr/bin/env python3
"""
Import Metro supplier purchase items from an Excel file.
Columns expected (case-insensitive): EAN, article number (or art number), price, name (or article name).
Unit and default_quantity are left null; you can fill them later in the app.

Usage:
  export SUPABASE_URL="https://xxx.supabase.co"
  export SUPABASE_SERVICE_ROLE_KEY="your-service-role-key"
  pip install openpyxl supabase
  python scripts/import_metro_items.py /path/to/metro_items_cleaned.xlsx

Optional: upload images (named by EAN) and set image_url:
  python scripts/import_metro_items.py /path/to/metro_items_cleaned.xlsx --images /path/to/metro_images

Images-only (no Excel, just upload images for existing Metro items by EAN):
  python scripts/import_metro_items.py /path/to/metro_items_cleaned.xlsx --images-only --images /path/to/metro_images
"""

import argparse
import os
import sys
from pathlib import Path

def _load_dotenv():
    """Load .env from project root (parent of scripts/) if present."""
    root = Path(__file__).resolve().parent.parent
    env_file = root / ".env"
    if not env_file.exists():
        return
    for line in env_file.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        k, v = k.strip(), v.strip().strip('"').strip("'")
        if k in ("SUPABASE_URL", "SUPABASE_SERVICE_ROLE_KEY", "SUPABASE_ANON_KEY") and k not in os.environ:
            os.environ[k] = v

def main():
    parser = argparse.ArgumentParser(description="Import Metro items from Excel into purchase_catalog_items")
    parser.add_argument("xlsx", nargs="?", help="Path to metro_items_cleaned.xlsx (optional when using --images-only)")
    parser.add_argument("--images", help="Folder of images named by EAN (e.g. 4337182011682.jpg)")
    parser.add_argument("--supplier-name", default="Metro", help="Supplier name to match (default: Metro)")
    parser.add_argument("--dry-run", action="store_true", help="Print rows without inserting")
    parser.add_argument("--update-prices", action="store_true", help="Only update last_known_price for existing items (by article number); no insert")
    parser.add_argument("--update-prices-by-ean", action="store_true", help="Only update last_known_price for existing items (match by EAN); no insert")
    parser.add_argument("--images-only", action="store_true", help="Only upload images from --images folder (match by EAN), no Excel import")
    args = parser.parse_args()

    _load_dotenv()
    try:
        import openpyxl
    except ImportError:
        print("Install: pip install openpyxl", file=sys.stderr)
        sys.exit(1)
    try:
        from supabase import create_client
        from postgrest.exceptions import APIError
    except ImportError:
        print("Install: pip install supabase", file=sys.stderr)
        sys.exit(1)

    url = (os.environ.get("SUPABASE_URL") or "").strip()
    key = (os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or os.environ.get("SUPABASE_ANON_KEY") or "").strip()
    if not url or not key:
        print("Set SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY (or add them to .env). Use the service_role key from Supabase Dashboard → Settings → API.", file=sys.stderr)
        sys.exit(1)

    xlsx_path = Path(args.xlsx or "")
    if not args.images_only:
        if not args.xlsx:
            print("Pass path to metro_items_cleaned.xlsx", file=sys.stderr)
            sys.exit(1)
        if not xlsx_path.exists():
            print(f"File not found: {xlsx_path}", file=sys.stderr)
            sys.exit(1)

    if args.images_only:
        if not args.images:
            print("With --images-only you must pass --images /path/to/metro_images", file=sys.stderr)
            sys.exit(1)
        supabase = create_client(url, key)
        try:
            try:
                supp = supabase.table("suppliers").select("id").ilike("name", f"%{args.supplier_name}%").execute()
            except Exception:
                supp = supabase.table("suppliers").select("id").filter("name", "ilike", f"%{args.supplier_name}%").execute()
        except APIError as e:
            err = str(e).lower()
            if "401" in err or "invalid api key" in err:
                print("API key rejected (401). Use the service_role key from Supabase Dashboard → Settings → API.", file=sys.stderr)
                print("Add to .env: SUPABASE_SERVICE_ROLE_KEY=your_service_role_key", file=sys.stderr)
                sys.exit(1)
            raise
        if not supp.data or len(supp.data) == 0:
            print(f"No supplier found matching '{args.supplier_name}'", file=sys.stderr)
            sys.exit(1)
        supplier_id = supp.data[0]["id"]
        print(f"Using supplier id: {supplier_id}")
        _run_image_upload(supabase, "purchase_items", supplier_id, args.images)
        return

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    # Normalize: lowercase, strip
    header_map = {h.lower().strip(): i for i, h in enumerate(headers) if h}
    print("Columns in sheet:", headers)

    def col(name_candidates):
        for n in name_candidates:
            n = n.lower().strip()
            if n in header_map:
                return header_map[n]
            for k in header_map:
                if n in k or k in n:
                    return header_map[k]
        return None

    i_ean = col(["ean", "ean number", "ean number "])
    i_art = col(["article number", "art number", "art. number", "article_number", "article no"])
    i_price = col(["price", "price per unit", "price (€)", "price(€)", "last_known_price"])
    i_name = col(["name", "article name", "article name ", "product name", "item name"])
    if i_name is None:
        print("Could not find a name column. Headers:", headers, file=sys.stderr)
        wb.close()
        sys.exit(1)
    if i_price is None:
        print("Could not find a price column. Headers:", headers, file=sys.stderr)
        wb.close()
        sys.exit(1)

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row) if row else []
        name = (row[i_name] if i_name is not None and i_name < len(row) else None) or ""
        if isinstance(name, (int, float)):
            name = str(int(name)) if isinstance(name, float) and name == int(name) else str(name)
        name = (name or "").strip()
        if not name:
            continue
        price_val = None
        if i_price is not None and i_price < len(row):
            p = row[i_price]
            if p is not None:
                if isinstance(p, (int, float)) and not isinstance(p, bool):
                    price_val = float(p)
                else:
                    s = str(p).strip().replace(",", ".").replace("€", "").replace(" ", "")
                    try:
                        price_val = float(s)
                    except (TypeError, ValueError):
                        pass
        article = ""
        if i_art is not None and i_art < len(row):
            a = row[i_art]
            if a is not None:
                article = str(int(a)) if isinstance(a, float) and a == int(a) else str(a).strip()
        ean_val = None
        if i_ean is not None and i_ean < len(row):
            e = row[i_ean]
            if e is not None:
                ean_val = str(int(e)) if isinstance(e, float) and e == int(e) else str(e).strip()
                if not ean_val or ean_val == "None":
                    ean_val = None
        rows.append({
            "name": name or "Unnamed",
            "article_number": article or None,
            "last_known_price": price_val,
            "ean": ean_val,
        })
    wb.close()

    print(f"Parsed {len(rows)} rows from Excel.")

    if args.dry_run:
        for i, r in enumerate(rows[:5]):
            print("  Sample:", r)
        if len(rows) > 5:
            print("  ...")
        return

    supabase = create_client(url, key)
    # Resolve supplier id by name (Metro)
    def fetch_supplier():
        try:
            return supabase.table("suppliers").select("id").ilike("name", f"%{args.supplier_name}%").execute()
        except Exception:
            return supabase.table("suppliers").select("id").filter("name", "ilike", f"%{args.supplier_name}%").execute()
    try:
        supp = fetch_supplier()
    except APIError as e:
        err = str(e).lower()
        if "401" in err or "invalid api key" in err:
            print("API key rejected (401). Use the service_role key from Supabase Dashboard → Settings → API (not the anon key).", file=sys.stderr)
            print("Add to .env: SUPABASE_SERVICE_ROLE_KEY=your_service_role_key", file=sys.stderr)
            sys.exit(1)
        raise
    if not supp.data or len(supp.data) == 0:
        print(f"No supplier found matching '{args.supplier_name}'", file=sys.stderr)
        sys.exit(1)
    supplier_id = supp.data[0]["id"]
    print(f"Using supplier id: {supplier_id}")

    if args.update_prices:
        updated = 0
        for r in rows:
            if not r.get("article_number") or r.get("last_known_price") is None:
                continue
            existing = supabase.table("purchase_catalog_items").select("id").eq("supplier_id", supplier_id).eq("article_number", r["article_number"]).execute()
            if existing.data and len(existing.data) > 0:
                supabase.table("purchase_catalog_items").update({"last_known_price": r["last_known_price"]}).eq("id", existing.data[0]["id"]).execute()
                updated += 1
        print(f"Updated last_known_price for {updated} items (by article number).")
        return

    if args.update_prices_by_ean:
        updated = 0
        for r in rows:
            ean_val = r.get("ean")
            price_val = r.get("last_known_price")
            if not ean_val or price_val is None:
                continue
            ean_str = str(ean_val).strip()
            existing = supabase.table("purchase_catalog_items").select("id").eq("supplier_id", supplier_id).eq("ean", ean_str).execute()
            if existing.data and len(existing.data) > 0:
                supabase.table("purchase_catalog_items").update({"last_known_price": price_val}).eq("id", existing.data[0]["id"]).execute()
                updated += 1
        print(f"Updated last_known_price for {updated} items (by EAN).")
        return

    bucket = "purchase_items"

    if args.images_only:
        _run_image_upload(supabase, bucket, supplier_id, args.images)
        return

    if args.images and not args.images_only:
        pass  # will run image upload after insert below
    inserted = 0
    skipped = 0
    for r in rows:
        if r.get("article_number"):
            existing = supabase.table("purchase_catalog_items").select("id").eq("supplier_id", supplier_id).eq("article_number", r["article_number"]).execute()
            if existing.data and len(existing.data) > 0:
                skipped += 1
                continue
        payload = {
            "supplier_id": supplier_id,
            "name": r["name"],
            "receipt_name": r.get("receipt_name") or r["name"],
            "article_number": r.get("article_number"),
            "ean": r.get("ean"),
            "last_known_price": r.get("last_known_price"),
            "unit": None,
            "default_quantity": None,
        }
        supabase.table("purchase_catalog_items").insert(payload).execute()
        inserted += 1

    print(f"Inserted {inserted} items, skipped {skipped} (duplicate article number).")

    if args.images:
        _run_image_upload(supabase, bucket, supplier_id, args.images)


def upload_image_to_supabase(supabase, bucket, image_path, ean, supplier_id):
    """Upload one image to Supabase Storage and return public URL. Replaces existing if upsert. EAN used for path."""
    try:
        path = Path(image_path)
        if not path.is_file():
            return None
        with open(path, "rb") as f:
            file_data = f.read()
        ext = path.suffix.lower()
        if ext not in (".jpg", ".jpeg", ".png", ".webp"):
            return None
        # Store under item_images/{supplier_id}/{ean}.jpg
        storage_path = f"item_images/{supplier_id}/{ean}.jpg"
        content_type = "image/jpeg" if ext in (".jpg", ".jpeg") else "image/png"
        file_options = {"content-type": content_type}
        supabase.storage.from_(bucket).upload(storage_path, file_data, file_options=file_options)
        public_url = supabase.storage.from_(bucket).get_public_url(storage_path)
        return public_url
    except Exception as e:
        print(f"  Upload error for {ean}: {e}", file=sys.stderr)
        return None


def _run_image_upload(supabase, bucket, supplier_id, images_dir_arg):
    """Upload images from folder (files named {ean}.jpg) and set purchase_catalog_items.image_url by EAN."""
    images_dir = Path(images_dir_arg)
    if not images_dir.is_dir():
        print(f"Images path is not a directory: {images_dir}", file=sys.stderr)
        return
    items = supabase.table("purchase_catalog_items").select("id, ean").eq("supplier_id", supplier_id).execute()
    ean_to_id = {str(it["ean"]): it["id"] for it in (items.data or []) if it.get("ean")}
    exts = {".jpg", ".jpeg", ".png", ".webp"}
    uploaded = 0
    for f in sorted(images_dir.iterdir()):
        if not f.is_file() or f.suffix.lower() not in exts:
            continue
        ean = f.stem
        if ean not in ean_to_id:
            continue
        image_url = upload_image_to_supabase(supabase, bucket, f, ean, supplier_id)
        if image_url:
            supabase.table("purchase_catalog_items").update({"image_url": image_url}).eq("id", ean_to_id[ean]).execute()
            uploaded += 1
    print(f"Uploaded {uploaded} images and set image_url in purchase_catalog_items.")


if __name__ == "__main__":
    main()
